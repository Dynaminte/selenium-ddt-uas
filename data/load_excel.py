# data/load_excel.py
"""
Utility: Baca data test dari file Excel (.xlsx)
Instalasi: pip install openpyxl
"""
import openpyxl


def load_excel(filename, sheet_name='Sheet1'):
    """
    Baca file Excel dari folder data/ dan kembalikan sebagai list of dict.
    Baris pertama dianggap sebagai header.
    """
    wb = openpyxl.load_workbook(f'data/{filename}')
    ws = wb[sheet_name]

    # Baris pertama = header
    headers = [cell.value for cell in ws[1]]

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip baris kosong
        if any(cell is not None for cell in row):
            data.append(dict(zip(headers, row)))

    return data


# Contoh penggunaan di test:
# from data.load_excel import load_excel
#
# @pytest.mark.parametrize('row', load_excel('login_data.xlsx'))
# def test_login_from_excel(self, driver, row):
#     page = LoginPage(driver)
#     page.login(row['username'], row['password'])
#     if row['expected'] == 'PASS':
#         assert page.is_login_successful()
#     else:
#         assert page.is_login_failed()
